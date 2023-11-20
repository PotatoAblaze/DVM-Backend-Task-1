from openpyxl import load_workbook, Workbook
from course import Course
from section import Section, DayOfTheWeek
from datetime import time
import sqlite3
import csv

class SectionPackage:
    
    def __init__(self):
        self.lecture = None
        self.tutorial = None
        self.lab = None

class Timetable:
    
    def __init__(self):
        self.all_courses = []
        self.timetable_courses = dict()
        
    def check_clashes(self, section_to_check : SectionPackage):
        
        for k in self.timetable_courses.keys():
            sp : SectionPackage = self.timetable_courses[k]
            do_dates_clash = False
            
            section_types = []
            if(sp.lecture != None):
                section_types.append(sp.lecture)
            if(sp.lab != None):
                section_types.append(sp.lab)
            if(sp.tutorial != None):
                section_types.append(sp.tutorial)
            
            for c in section_types:
                for day in c.days:
                    if(section_to_check.days.count(day) > 0):
                        do_dates_clash = True
                
                if(do_dates_clash):
                    start_a = section_to_check.start_time.hour
                    end_a = section_to_check.start_time.hour + section_to_check.duration
                    start_b = c.start_time.hour
                    end_b = c.start_time.hour + c.duration
                    
                    if(not (end_a <= start_b or end_b <= start_a)):
                        return(k)
        
        return(False)
                        
                
        
    def load_courses_into_database(self):
        workbook = load_workbook('Courses.xlsx')
        sheet = workbook.active
        
        database_connection = sqlite3.connect('courses.db')
        
        cursor = database_connection.cursor()
        
        columns = [cell.value for cell in sheet[1]]
        delete_old_table_query = f"DROP TABLE IF EXISTS courses_table"
        cursor.execute(delete_old_table_query)
        create_table_query = f"CREATE TABLE courses_table ({', '.join([ f'{col} TEXT' for col in columns])})"
        cursor.execute(create_table_query)
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            insert_query = f"INSERT INTO courses_table ({', '.join(columns)}) VALUES ({', '.join(['?' for _ in columns])})"
            cursor.execute(insert_query, row)
        
        database_connection.commit()
        database_connection.close()
        
    
    def display_database(self):
        database_connection = sqlite3.connect('courses.db')
        cursor = database_connection.cursor()
        
        cursor.execute("SELECT * FROM courses_table")
        
        rows = cursor.fetchall()

        for row in rows:
            print(row)
            
        database_connection.close()
        
    def display_courses_in_memory(self):
        for course_entry in self.all_courses:
            print(course_entry)
            print("\n")
            course_entry.get_all_sections()
    
    def display_courses_in_memory_without_sections(self):
        for course_entry in self.all_courses:
            print(course_entry)
            print()
    
    def export_to_csv(self):
        workbook = Workbook()
        workbook.create_sheet("Timetable")
        
        timetable_sheet = workbook['Timetable']
        
        for x in range(1, 8):
            timetable_sheet.cell(column=1, row=x+1, value=DayOfTheWeek(x).name)
        
        for x in range(2, 14):
            time_string = f"{x + 6}:00 to {x+7}:00"
            timetable_sheet.cell(row=1, column=x, value=time_string)
        
        for course_package_key in self.timetable_courses.keys():
            
            course_package = self.timetable_courses[course_package_key]
            
            section_types = []
            if(course_package.lecture != None):
                section_types.append(course_package.lecture)
            if(course_package.lab != None):
                section_types.append(course_package.lab)
            if(course_package.tutorial != None):
                section_types.append(course_package.tutorial)
            
            course_object = None
            for crse in self.all_courses:
                if(crse.course_id == course_package_key):
                    course_object = crse
                    break
            
            for sec in section_types:
                for day in sec.days:
                    for hour in range(sec.start_time.hour - 7, sec.start_time.hour + sec.duration - 7):
                        cell_value = f"{sec.section_number} {course_object.course_name} ({sec.room_no})"
                        timetable_sheet.cell(row=day+1, column=hour+1, value=cell_value)
        
        with open("timetable.csv", "w", newline="") as output_file:
            csv_writer = csv.writer(output_file)
            for row in timetable_sheet.iter_rows(values_only=True):
                csv_writer.writerow(row)
            
    
    
    def load_courses_from_database(self):
        database_connection = sqlite3.connect('courses.db')
        cursor = database_connection.cursor()
        
        cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='courses_table'")
    
        result = cursor.fetchone()

        if(result == None):
            database_connection.close()
            return False
        
        cursor.execute("SELECT * FROM courses_table")
        
        rows = cursor.fetchall()
        current_course = None
        for row in rows:
            if(row[0] != None):
                current_course = Course()
                current_course.course_id = row[0]
                current_course.course_name = row[1]
                current_course.units = int(float(row[2]))
                current_course.lecture_hours = int(float(row[3]))
                current_course.lab_hours = int(float(row[4]))
                self.all_courses.append(current_course)
            
            new_section = Section()
            new_section.section_number = row[5]
            new_section.room_no = int(float(row[7]))
            
            instructor_names = row[6].strip().split('\n')
            instructor_names = [name.strip() for name in instructor_names]
            new_section.instructors = instructor_names
            
            new_section.start_time = time(hour=int(float(row[8])), minute=0, second=0)
            
            new_section.duration = int(float(row[9]))
            
            class_days = row[10].strip().split()
            
            for day in class_days:
                day_int = 0
                match day:
                    case 'M':
                        day_int = 1
                    case 'T':
                        day_int = 2
                    case 'W':
                        day_int = 3
                    case 'Th':
                        day_int = 4
                    case 'F':
                        day_int = 5
                    case 'S':
                        day_int = 6
                    case 'Su':
                        day_int = 7
                
                if(day_int != 0):
                    new_section.days.append(day_int)
            
            current_course.sections.append(new_section)
            
        database_connection.close()
        return(True)
        
        
    
    